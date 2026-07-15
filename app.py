"""
Consolidated Resource Report Generator — Streamlit App
--------------------------------------------------------
Upload:
  - MIS Report (.xlsx)              ← weekly input
  - Employee.xlsx (optional)        ← uses bundled file if not uploaded
  - Project.xlsx  (optional)        ← uses bundled file if not uploaded

Click "Generate Report" to download the consolidated output.
"""

import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import streamlit as st

# ──────────────────────────────────────────────
# CONFIGURATION
# ──────────────────────────────────────────────

EXCLUDED_DEPARTMENTS = {
    'Administration',
    'Collaboration Services',
    'Digital Marketing',
    'Finance',
    'HR',
    'IT System Engineering',
    'Maintenance',
    'Management',
    'Products and Services',
    'Salesforce',
}

KEYWORD_RULES = [
    ('highsystem', 'Nexus Schweiz AG'),
    ('star',       'STAR Enterprises AG'),
    ('sky',        'STAR Enterprises AG'),
    ('qiki',       'QiKi Technologics AG'),
    ('mawi',       'Nexus Schweiz AG'),
    ('vsm',        'Prozessraum AG'),
    ('design quest', 'Bix Bytes Solutions AG'),
    ('coco grid',  'COCOGRID'),
    ('nexus',      'Nexus Schweiz AG'),
]

# Design cloned from Resource_Allocation_-_2026.xlsx → Master Sheet
FONT_NAME     = 'Aptos Narrow'
HEADER_FILL   = 'A6CAEC'
FLAG_FILL     = 'FFFF00'
HEADERS       = ['Department', 'Name', 'Role', 'Technology', 'Client',
                  'Project', 'Allocation', 'FTE', 'Billable (Y/N)', 'Location', 'Notes']
COLUMN_WIDTHS = [14.86, 31.43, 20.86, 22.57, 24.71, 32.29, 13.14, 6.71, 15.71, 11.71, 15.43]


# ──────────────────────────────────────────────
# CORE LOGIC  (no file-system paths — everything in-memory)
# ──────────────────────────────────────────────

def load_employee_map(file_obj):
    df = pd.read_excel(file_obj, header=None)
    role_map, location_map = {}, {}
    for i, row in df.iterrows():
        vals = [str(v).strip() if pd.notna(v) else '' for v in row]
        if 'Employee Name' in vals and 'Role' in vals:
            ec = vals.index('Employee Name')
            rc = vals.index('Role')
            lc = vals.index('Location') if 'Location' in vals else None
            for _, dr in df.iloc[i + 1:].iterrows():
                emp  = str(dr.iloc[ec]).strip() if pd.notna(dr.iloc[ec]) else ''
                role = str(dr.iloc[rc]).strip() if pd.notna(dr.iloc[rc]) else ''
                loc  = str(dr.iloc[lc]).strip() if (lc is not None and pd.notna(dr.iloc[lc])) else ''
                if emp and emp.lower() != 'nan':
                    role_map[emp.lower()]     = role
                    location_map[emp.lower()] = loc
            break
    return role_map, location_map


def load_client_map(file_obj):
    df = pd.read_excel(file_obj, header=None)
    client_map = {}
    for i, row in df.iterrows():
        vals = [str(v).strip() if pd.notna(v) else '' for v in row]
        if 'Project Name' in vals and 'Client' in vals:
            pc = vals.index('Project Name')
            cc = vals.index('Client')
            for _, dr in df.iloc[i + 1:].iterrows():
                proj   = str(dr.iloc[pc]).strip() if pd.notna(dr.iloc[pc]) else ''
                client = str(dr.iloc[cc]).strip() if pd.notna(dr.iloc[cc]) else ''
                if proj and proj.lower() != 'nan':
                    client_map[proj.lower()] = client
            break
    return client_map


def lookup_client(proj_name, client_map):
    key = proj_name.lower()
    if key in client_map:
        return client_map[key]
    for map_key, client in client_map.items():
        if map_key in key or key in map_key:
            return client
    for keyword, client in KEYWORD_RULES:
        if keyword in key:
            return client
    return ''


def identify_departments(df):
    not_null_idx = df[df[0].notna()].index.tolist()
    dept_set = set()
    for i, idx in enumerate(not_null_idx):
        val = df.iloc[idx][0]
        if val in ['MIS Report', 'Employees']:
            continue
        next_idx = not_null_idx[i + 1] if i + 1 < len(not_null_idx) else None
        if next_idx and next_idx == idx + 1:
            dept_set.add(val)
    return dept_set


def parse_data(df, dept_set):
    current_dept, current_emp = None, None
    emp_to_record = {}
    for idx in range(len(df)):
        row = df.iloc[idx]
        col0, col1, col2, col3 = row[0], row[1], row[2], row[3]
        if pd.notna(col0):
            val = col0
            if val in ['MIS Report', 'Employees']:
                continue
            if val in dept_set:
                current_dept, current_emp = val, None
            else:
                current_emp = val
                if current_dept and current_dept not in EXCLUDED_DEPARTMENTS:
                    key = (current_dept, current_emp)
                    if key not in emp_to_record:
                        emp_to_record[key] = {
                            'dept': current_dept, 'employee': current_emp,
                            'pr_projects': set(), 'ipr_projects': set(),
                        }
        elif pd.notna(col1) and pd.notna(col2) and pd.notna(col3):
            proj_num  = str(col2).strip()
            proj_name = str(col3).strip()
            if current_emp and current_dept and current_dept not in EXCLUDED_DEPARTMENTS:
                key = (current_dept, current_emp)
                if key in emp_to_record:
                    if proj_num.startswith('PR'):
                        emp_to_record[key]['pr_projects'].add((proj_num, proj_name))
                    elif proj_num.startswith('IPR'):
                        emp_to_record[key]['ipr_projects'].add((proj_num, proj_name))
    return emp_to_record


def build_output_rows(emp_to_record, role_map, location_map, client_map):
    output_rows = []
    for (dept, emp), rec in emp_to_record.items():
        pr  = sorted(rec['pr_projects'])
        ipr = sorted(rec['ipr_projects'])
        if pr:
            projects = [{'name': p[1], 'billable': 'Yes'} for p in pr]
        elif ipr:
            projects = [{'name': ipr[0][1], 'billable': 'No'}]
        else:
            continue
        allocation = 'Part Time' if len(projects) > 1 else 'Full Time'
        role       = role_map.get(emp.lower(), '')
        location   = location_map.get(emp.lower(), '')
        for proj in projects:
            client = lookup_client(proj['name'], client_map)
            output_rows.append({
                'Department': dept, 'Name': emp, 'Role': role,
                'Technology': '', 'Client': client, 'Project': proj['name'],
                'Allocation': allocation, 'FTE': '',
                'Billable (Y/N)': proj['billable'], 'Location': location, 'Notes': '',
            })
    output_rows.sort(key=lambda x: (x['Department'], x['Name'], x['Project']))
    return output_rows


def build_excel(output_rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Sheet"

    hfill       = PatternFill('solid', start_color=HEADER_FILL)
    hfont       = Font(bold=True, name=FONT_NAME, size=12)
    dfont       = Font(name=FONT_NAME, size=11)
    flag_fill   = PatternFill('solid', start_color=FLAG_FILL)
    thin        = Side(style='thin')
    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell_align  = Alignment(horizontal='left', vertical='top')

    for col, h in enumerate(HEADERS, 1):
        cell           = ws.cell(row=1, column=col, value=h)
        cell.font      = hfont
        cell.fill      = hfill
        cell.alignment = cell_align
        cell.border    = thin_border

    ws.row_dimensions[1].height = 15.75

    for row_num, r in enumerate(output_rows, start=2):
        for col, h in enumerate(HEADERS, 1):
            val  = r[h]
            cell = ws.cell(row=row_num, column=col, value=val if val != '' else None)
            cell.font      = dfont
            cell.alignment = cell_align
            cell.border    = thin_border
            if h == 'Role':
                cell.fill = flag_fill
            elif h == 'FTE':
                cell.fill = flag_fill
            elif h == 'Location' and not val:
                cell.fill = flag_fill

    for letter, width in zip('ABCDEFGHIJK', COLUMN_WIDTHS):
        ws.column_dimensions[letter].width = width

    ws.auto_filter.ref = f"A1:K{len(output_rows) + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ──────────────────────────────────────────────
# STREAMLIT UI
# ──────────────────────────────────────────────

st.set_page_config(
    page_title="Consolidated Resource Report",
    page_icon="📊",
    layout="centered",
)

st.title("📊 Consolidated Resource Report Generator")
st.caption("Upload your MIS Report to generate a formatted consolidated report matching the Resource Allocation Master Sheet design.")

st.divider()

col1, col2 = st.columns(2)

with col1:
    st.subheader("📁 Required")
    mis_file = st.file_uploader("MIS Report (.xlsx)", type=["xlsx"], key="mis")

with col2:
    st.subheader("📂 Optional overrides")
    emp_file  = st.file_uploader("Employee.xlsx", type=["xlsx"], key="emp",
                                  help="Leave blank to use the bundled Employee.xlsx")
    proj_file = st.file_uploader("Project.xlsx",  type=["xlsx"], key="proj",
                                  help="Leave blank to use the bundled Project.xlsx")

st.divider()

if mis_file:
    if st.button("⚙️ Generate Report", type="primary", use_container_width=True):
        with st.spinner("Processing..."):
            try:
                # Load mapping files — prefer upload, fall back to bundled
                import os
                BASE = os.path.dirname(os.path.abspath(__file__))

                emp_source  = emp_file  if emp_file  else os.path.join(BASE, "Employee.xlsx")
                proj_source = proj_file if proj_file else os.path.join(BASE, "Project.xlsx")

                role_map, location_map = load_employee_map(emp_source)
                client_map             = load_client_map(proj_source)

                # Parse MIS report
                df = pd.read_excel(mis_file, header=None, sheet_name=0)
                dept_set      = identify_departments(df)
                emp_to_record = parse_data(df, dept_set)
                output_rows   = build_output_rows(emp_to_record, role_map, location_map, client_map)

                if not output_rows:
                    st.error("No matching data found in the MIS Report. Please check the file.")
                else:
                    excel_buf = build_excel(output_rows)

                    # Summary metrics
                    dept_count = len(set(r['Department'] for r in output_rows))
                    emp_count  = len(set(r['Name']       for r in output_rows))
                    m1, m2, m3 = st.columns(3)
                    m1.metric("Total Rows",    len(output_rows))
                    m2.metric("Departments",   dept_count)
                    m3.metric("Employees",     emp_count)

                    # Derive output filename from MIS filename
                    base_name   = os.path.splitext(mis_file.name)[0]
                    output_name = f"Consolidated_{base_name}.xlsx"

                    st.success(f"✅ Report generated — {len(output_rows)} rows across {dept_count} departments.")

                    st.download_button(
                        label="⬇️ Download Consolidated Report",
                        data=excel_buf,
                        file_name=output_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                    )

            except Exception as e:
                st.error(f"Something went wrong: {e}")
                st.exception(e)
else:
    st.info("👆 Upload your MIS Report (.xlsx) to get started.")
